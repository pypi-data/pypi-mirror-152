import os

import openpyxl as xl
from openpyxl.utils import get_column_letter

from miroflowexport.internal import excelcolumns

GRANULARITY_WEEKS = "weeks"
GRANULARITY_DAYS = "days"



def sanitize_version(log, version):
    try:
        sanitized = [
            int(v)
            for v in version
        ]
        return sanitized
    except:
        log.error("Interpreting version info '{}' failed, I continue without version information.")
        
    return None

def find_column_for_cell(log, ws, row, cell_value, max_cols_to_search = 30):
    for col in range(1, max_cols_to_search):
        col_letter = get_column_letter(col)
        if ws["{}{}".format(col_letter, row)].value == cell_value:
            return col_letter
            
    return None

def metadata_from_excel(log, path):
    log.debug("Trying to load metadata from '{}' ...".format(path))
    date_offset = None
    version = None
    if not os.path.exists(path):
        log.error("Cannot find Excel file '{}' to read from. I skip loading from Excel".format(path))
        return date_offset, version

    try:
        wb = xl.load_workbook(path)
    except:
        log.exception("Loading Excel from '{}' failed. I skip import.".format(path))
        return date_offset, version

    if not excelcolumns.DEFAULT_SHEET_META in wb.sheetnames:
        log.error("There is not worksheet called '{}' in file '{}'. I stop importing meta information.".format(excelcolumns.DEFAULT_SHEET_META, path))
        return date_offset, version

    ws = wb[excelcolumns.DEFAULT_SHEET_META]
    version_info_found = True
    version = []
    for colnum, colheader in [
        (excelcolumns.COLNUM_META_VERSION_MAJOR, excelcolumns.COL_META_VERSION_MAJOR),
        (excelcolumns.COLNUM_META_VERSION_MINOR, excelcolumns.COL_META_VERSION_MINOR),
        (excelcolumns.COLNUM_META_VERSION_PATCH, excelcolumns.COL_META_VERSION_PATCH),
    ]:
        found = ws.cell(column = colnum, row = excelcolumns.ROWNUM_META_INFO - 1).value == colheader
        version_info_found &= found
        if not found:
            log.error("Cannot find version keyword '{}' in cell ({}, {})".format(
                colheader,
                get_column_letter(colnum),
                excelcolumns.ROWNUM_META_INFO - 1,
            ))
        else:
            version += [ws.cell(column = colnum, row = excelcolumns.ROWNUM_META_INFO).value]
    log.debug("Found version info {} in Excel metadata sheet.".format(version))

    if not ws.cell(column = excelcolumns.COLNUM_META_DATE_OFFSET, row = excelcolumns.ROWNUM_META_INFO - 1).value == excelcolumns.COL_META_DATE_OFFSET:
        log.error("Cannot found date offset keywork '{}' in cell ({}, {}).".format(
            excelcolumns.COL_META_DATE_OFFSET,
            get_column_letter(excelcolumns.COLNUM_META_DATE_OFFSET),
            excelcolumns.ROWNUM_META_INFO - 1,
        ))
    else:
        date_offset = ws.cell(column = excelcolumns.COLNUM_META_DATE_OFFSET, row = excelcolumns.ROWNUM_META_INFO).value
        log.debug("Found date offset '{}' in Excel.".format(date_offset))

    return date_offset, sanitize_version(log, version)
