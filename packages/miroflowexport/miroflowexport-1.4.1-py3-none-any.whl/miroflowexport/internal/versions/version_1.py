import datetime
import math
import os

import openpyxl as xl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.differential import DifferentialStyle

from miroflowexport.internal import excelcolumns
from miroflowexport.internal import task
from miroflowexport import version
from miroflowexport.internal.excelcolumns import *
from miroflowexport.internal.versions import common

BG_COLOR_HORIZON = 'E2E5DE'
BG_COLOR_PLAN = '1AA7EC'
BG_COLOR_WEEK_HIGHLIGHT = 'E9CFEC'
BG_COLOR_WEEKEND_HIGHLIGHT = 'D3D3D3'

COLOR_PROGRESS_0_NOT_STARTED = '1AA7EC'
COLOR_PROGRESS_1_STARTED = 'FCF4A3'
COLOR_PROGRESS_2_WORKING = 'F8E473'
COLOR_PROGRESS_3_NEARLY_DONE = '80b280'
COLOR_PROGRESS_4_DONE = '33b333'

TASK_FILL_HORIZON = 'o'
TASK_FILL_PLAN = 'x'

HEADER_ROW = 1
DATA_ROW_START = 4

THRESHOLD_FOLD_FUTURE_TASKS = 10

def apply_conditional_progress_coloring(ws, base_formula_plan, formula_range, color, reference_column, upper_threshold):
    color_fill_plan = PatternFill(bgColor = color)
    style_plan = DifferentialStyle(fill = color_fill_plan, font = Font(color = color))
    rule_plan = Rule(type="expression", dxf=style_plan, stopIfTrue=True)
    rule_plan.formula = ["AND({}, {})".format(base_formula_plan, "{}<={}".format(reference_column, upper_threshold))]
    ws.conditional_formatting.add(formula_range, rule_plan)

def apply_conditional_current_week_coloring(log, ws, color, formula_range, reference_column, reference_row):
    next_column = get_column_letter(column_index_from_string(reference_column) + 1)
    color_fill_cell = PatternFill(bgColor = color)
    style_cell = DifferentialStyle(fill = color_fill_cell)
    rule_cell = Rule(type="expression", dxf = style_cell, stopIfTrue = False)
    formula = "AND({refcol}${refrow}<=TODAY(), {nextcol}${refrow}>TODAY())".format(
            refcol = reference_column,
            nextcol = next_column,
            refrow = reference_row,
        )
    log.debug("Applying conditional week highlight for range {} formatting using formula: {}".format(formula_range, formula))
    rule_cell.formula = [formula]
    ws.conditional_formatting.add(formula_range, rule_cell)

def apply_weekend_coloring(log, ws, formula_range, color, reference_column, reference_row):
    color_fill = PatternFill(bgColor = color)
    style = DifferentialStyle(fill = color_fill, font = Font(color = color))
    rule = Rule(type = "expression", dxf = style, stopIfTrue = True)
    rule.formula = [
        "OR(WEEKDAY({refcol}${refrow}) = 1, WEEKDAY({refcol}${refrow}) = 7)".format(
            refcol = reference_column,
            refrow = reference_row
        )
    ]
    ws.conditional_formatting.add(formula_range, rule)

def sort_by_critical_path(log, task_list):
    crit_only = [task for task in task_list if task.is_on_critical_path()]
    non_crit = [task for task in task_list if not task.is_on_critical_path()]
    crit_only = sorted(crit_only, key = lambda t: t.start())
    non_crit = sorted(non_crit, key = lambda t: t.earliest_start())

    crit_only.extend(non_crit)
    return crit_only

def get_calendar_time_for_worktime(worktime_end, horizon_granularity):
    if horizon_granularity == common.GRANULARITY_DAYS:
        workdays_per_week = 5
        full_weeks = math.floor(worktime_end / workdays_per_week)
        remaining_days = worktime_end % workdays_per_week
        caltime = full_weeks * 7 + remaining_days
        if remaining_days == 0:
            # do not count the weekend if the last day was a Friday
            caltime -= 2

        return caltime

    return worktime_end

def get_caltime_passed_since_offset(date_offset, horizon_granularity, date_today = datetime.date.today()):
    if isinstance(date_offset, datetime.datetime):
        date_offset = date_offset.date()

    if horizon_granularity == common.GRANULARITY_DAYS:
        delta = (date_today - date_offset).days
        return delta

    _, week_offset, _ = date_offset.isocalendar()
    _, week_today, _ = date_today.isocalendar()
    weeks_passed = week_today - week_offset
    return weeks_passed

def get_caltime_from_worktime_formula(col, row):
    caltime_from_worktime_formula = 'ROUNDDOWN({col}{row}/5, 0) * 7 + MOD({col}{row}, 5)'.format(col = col, row = row)    
    return caltime_from_worktime_formula

def get_plan_formula(
    horizon_granularity,
    col_letter_task_start,
    col_letter_task_finish,
    col_letter_task_es,
    col_letter_task_lf,
    current_row,
    caltime,
    fill_plan = TASK_FILL_PLAN,
    fill_horizon = TASK_FILL_HORIZON,
):
    if horizon_granularity == common.GRANULARITY_DAYS:
        formula = '=IF('\
                    "AND("\
                        "{t_start}<={caltime},"\
                        "{t_end}>={caltime}"\
                    "),"\
                    '"{fill_plan}",'\
                    "IF("\
                        "AND("\
                            "{t_es}<={caltime},"\
                            "{t_lf}>={caltime}"\
                        "),"\
                        '"{fill_horizon}",'\
                        '""'\
                    ")"\
                ")".format(
                    t_start = get_caltime_from_worktime_formula(col_letter_task_start, current_row),
                    t_end = get_caltime_from_worktime_formula(col_letter_task_finish, current_row),
                    t_es = get_caltime_from_worktime_formula(col_letter_task_es, current_row),
                    t_lf = get_caltime_from_worktime_formula(col_letter_task_lf, current_row),
                    caltime = caltime,
                    fill_plan = fill_plan,
                    fill_horizon = fill_horizon,
                )
        return formula

    formula = '=IF('\
                "AND("\
                    "{col_letter_task_start}{row_num}<={caltime},"\
                    "{col_letter_task_finish}{row_num}>={caltime}"\
                "),"\
                '"{fill_plan}",'\
                "IF("\
                    "AND("\
                        "{col_letter_task_es}{row_num}<={caltime},"\
                        "{col_letter_task_lf}{row_num}>={caltime}"\
                    "),"\
                    '"{fill_horizon}",'\
                    '""'\
                ")"\
            ")".format(
                col_letter_task_start = col_letter_task_start,
                col_letter_task_finish = col_letter_task_finish,
                col_letter_task_es = col_letter_task_es,
                col_letter_task_lf = col_letter_task_lf,
                row_num = current_row,
                caltime = caltime,
                fill_plan = fill_plan,
                fill_horizon = fill_horizon,
            )
    return formula

def __tasks_to_excel_tasks(log, wb, tasks_dict, sheet_tasks, date_offset, horizon_granularity, colors):
    ws = wb.active
    ws.title = sheet_tasks

    caltime_passed = get_caltime_passed_since_offset(date_offset, horizon_granularity)
    log.debug("Calendar time passed since date offset: {} {}".format(caltime_passed, horizon_granularity))

    worktime_start = 0
    caltime_end = max([
        task.latest_finish()
        for task in tasks_dict.values()
    ])
    caltime_end = get_calendar_time_for_worktime(caltime_end, horizon_granularity)

    cols_caltime = [
        "{}{}".format(EXCEL_COL_NAMES[COL_TASK_WEEK_PREFIX], cal)
        for cal in range(caltime_end + 1)
    ]

    header_row = [
        EXCEL_COL_NAMES[COL_TASK_ID],
        EXCEL_COL_NAMES[COL_TASK_NAME],
        EXCEL_COL_NAMES[COL_TASK_START],
        EXCEL_COL_NAMES[COL_TASK_ES],
        EXCEL_COL_NAMES[COL_TASK_LF],
        EXCEL_COL_NAMES[COL_TASK_EFFORT],
        EXCEL_COL_NAMES[COL_HORIZON_START],
        EXCEL_COL_NAMES[COL_HORIZON_END],
        EXCEL_COL_NAMES[COL_TASK_FINISH],
        EXCEL_COL_NAMES[COL_TASK_COSTS],
        EXCEL_COL_NAMES[COL_TASK_PROGRESS],
    ]
    date_row = [
        ""
        for _ in header_row
    ]
    caltime_row = [
        header
        for header in header_row
    ]
    col_num_progress = header_row.index(EXCEL_COL_NAMES[COL_TASK_PROGRESS]) + 1
    col_num_costs = header_row.index(EXCEL_COL_NAMES[COL_TASK_COSTS]) + 1

    col_letter_task_start = get_column_letter(header_row.index(EXCEL_COL_NAMES[COL_TASK_START]) + 1)
    col_letter_task_effort = get_column_letter(header_row.index(EXCEL_COL_NAMES[COL_TASK_EFFORT]) +1)
    col_letter_task_finish = get_column_letter(header_row.index(EXCEL_COL_NAMES[COL_TASK_FINISH]) + 1)
    col_letter_task_es = get_column_letter(header_row.index(EXCEL_COL_NAMES[COL_TASK_ES]) + 1)
    col_letter_task_lf = get_column_letter(header_row.index(EXCEL_COL_NAMES[COL_TASK_LF]) + 1)

    col_letter_progress = get_column_letter(col_num_progress)
    col_letter_task_costs = get_column_letter(col_num_costs)
    col_letter_task_name = get_column_letter(header_row.index(EXCEL_COL_NAMES[COL_TASK_NAME]) + 1)

    groups_to_hide = [
        # Hide Task ID
        (get_column_letter(header_row.index(EXCEL_COL_NAMES[COL_TASK_ID]) + 1), get_column_letter(header_row.index(EXCEL_COL_NAMES[COL_TASK_ID]) + 1)),
        # Hide suporting information
        (get_column_letter(header_row.index(EXCEL_COL_NAMES[COL_TASK_ES]) + 1), get_column_letter(header_row.index(EXCEL_COL_NAMES[COL_TASK_FINISH]) + 1)),
    ]

    num_cols_before_horizon = len(header_row)
    header_row.extend(cols_caltime)

    col_letters_for_horizon_cols = [
        get_column_letter(hzn_col)
        for hzn_col in range(num_cols_before_horizon + 1, num_cols_before_horizon + 1 + len(cols_caltime))
    ]

    date_row_num = 2
    date_offset_reference = "{metasheet}!{colletter}{row}".format(
        metasheet = excelcolumns.DEFAULT_SHEET_META,
        colletter = get_column_letter(excelcolumns.COLNUM_META_DATE_OFFSET),
        row = excelcolumns.ROWNUM_META_INFO,
    )
    date_horizon_start = date_offset_reference
    if horizon_granularity == common.GRANULARITY_DAYS:
        date_horizon_start = "{start} + MOD(7 - (WEEKDAY({start}) - 2), 7)".format(
            start = date_offset_reference,
        )

    formula_date_offset = "={ref}".format(
        ref = date_horizon_start,
    )
    caltime_step = 7
    if horizon_granularity == common.GRANULARITY_DAYS:
        caltime_step = 1

    date_row.append(formula_date_offset)
    date_row.extend([
        "={col}{row}+{caltime_step}".format(col = col_letter, row = date_row_num, caltime_step = caltime_step)
        for col_letter in col_letters_for_horizon_cols[:-1]
    ])

    caltime_row_num = 3
    caltime_row.extend([
        "=WEEKNUM({col}{row})".format(col = col_letter, row = date_row_num)
        for col_letter in col_letters_for_horizon_cols
    ])

    ws.append(header_row)
    ws.append(date_row)
    for col_num in range(num_cols_before_horizon + 1, num_cols_before_horizon + 1 + len(cols_caltime)):
        ws.cell(column = col_num, row = date_row_num).number_format = "yyyy-mm-dd"
        ws.cell(column = col_num, row = date_row_num).font = Font(size = 8)
        ws.cell(column = col_num, row = date_row_num).alignment = Alignment(text_rotation = 90)
    ws.append(caltime_row)
    for col_num in range(1, num_cols_before_horizon + 1 + len(cols_caltime)):
        ws.cell(column = col_num, row = caltime_row_num).font = Font(b = True)

    data_row_start = DATA_ROW_START

    rows_to_hide = [
        (1, date_row_num - 1)
    ]

    current_row = data_row_start - 1
    wrote_progress_formula_to_log = False
    row_fold_future_started = False
    row_far_far_away_start = None
    row_far_far_away_end = None

    threshold = THRESHOLD_FOLD_FUTURE_TASKS + caltime_passed

    for task in sort_by_critical_path(log, list(tasks_dict.values())):
        current_row += 1
        formula_task_end = "={col_letter_start}{row_num}+{col_letter_effort}{row_num}-1".format(
            col_letter_start = col_letter_task_start,
            col_letter_effort = col_letter_task_effort,
            row_num = current_row,
        )
        content_row = [
            task.id(),
            task.name(),
            task.start(),
            task.earliest_start(),
            task.latest_finish(),
            task.effort(),
            worktime_start,
            caltime_end,
            formula_task_end,
            task.costs(),
            task.progress(),
        ]
        content_row.extend([
            get_plan_formula(
                horizon_granularity = horizon_granularity,
                col_letter_task_start = col_letter_task_start,
                col_letter_task_finish = col_letter_task_finish,
                col_letter_task_es = col_letter_task_es,
                col_letter_task_lf = col_letter_task_lf,
                current_row = current_row,
                caltime = caltime,
                fill_plan = TASK_FILL_PLAN,
                fill_horizon = TASK_FILL_HORIZON,
            )
            # "" if week < task.earliest_start() or week > task.latest_finish() else TASK_FILL_PLAN if week >= task.start() and week <= task.end() else TASK_FILL_HORIZON
            for caltime in range(worktime_start, caltime_end + 1)
        ])
        if not wrote_progress_formula_to_log:
            log.debug("Progress formula: {}".format(content_row[-1]))
            wrote_progress_formula_to_log = True
        ws.append(content_row)
        ws["{}{}".format(col_letter_progress, current_row)].number_format = "0%"

        if colors:
            task_name_color = task.get_color()
            if task_name_color:
                ws["{}{}".format(col_letter_task_name, current_row)].fill = PatternFill(
                    patternType = "solid",
                    fgColor = task.get_color().replace("#", "")
                )

        if not row_far_far_away_start:
            if task.earliest_start() > threshold:
                row_far_far_away_start = current_row
                
        if row_far_far_away_start and not row_far_far_away_end:
            if task.earliest_start() <= threshold:
                row_far_far_away_end = current_row - 1

    if row_far_far_away_start and not row_far_far_away_end:
        row_far_far_away_end = current_row

    if row_far_far_away_start:
        rows_to_hide += [
            (row_far_far_away_start, row_far_far_away_end)
        ]

    formula_base = "{}{}=".format(col_letters_for_horizon_cols[0], data_row_start)
    log.debug("Formula base for conditional formatting: {}".format(formula_base))

    formula_range = "{}{}:{}{}".format(
        col_letters_for_horizon_cols[0],
        data_row_start,
        col_letters_for_horizon_cols[-1],
        len(tasks_dict.keys()) + data_row_start
    )
    log.debug("Formula range for conditional formatting: {}".format(formula_range))

    formula_horizon = '{}"{}"'.format(formula_base, TASK_FILL_HORIZON)
    formula_plan = '{}"{}"'.format(formula_base, TASK_FILL_PLAN)
    log.debug("Formula for horizon formatting: {}".format(formula_horizon))
    log.debug("Formula for plan formatting:    {}".format(formula_plan))

    ref_column_for_week_highlight = col_letters_for_horizon_cols[0]
    ref_row_for_week_highlight = date_row_num
    formula_range_for_week_highlight = "{}{}:{}{}".format(
        col_letters_for_horizon_cols[0],
        date_row_num,
        col_letters_for_horizon_cols[-1],
        len(tasks_dict.keys()) + data_row_start - 1
    )

    if horizon_granularity == common.GRANULARITY_DAYS:
        apply_weekend_coloring(
            log = log,
            ws = ws,
            color = BG_COLOR_WEEKEND_HIGHLIGHT,
            formula_range = formula_range_for_week_highlight,
            reference_column = ref_column_for_week_highlight,
            reference_row = ref_row_for_week_highlight,
        )

    font_horizon = Font(color = BG_COLOR_HORIZON)
    
    color_fill_horizon = PatternFill(bgColor = BG_COLOR_HORIZON)
    style_horizon = DifferentialStyle(fill = color_fill_horizon, font = font_horizon)
    rule_horizon = Rule(type="expression", dxf=style_horizon, stopIfTrue=True)
    rule_horizon.formula = [formula_horizon]
    ws.conditional_formatting.add(formula_range, rule_horizon)

    ref_column = "${}{}".format(col_letter_progress, data_row_start)
    for color, threshold in [
        (COLOR_PROGRESS_0_NOT_STARTED, 0),
        (COLOR_PROGRESS_1_STARTED, 0.25),
        (COLOR_PROGRESS_2_WORKING, 0.75),
        (COLOR_PROGRESS_3_NEARLY_DONE, 0.99),
        (COLOR_PROGRESS_4_DONE, 1),
    ]:
        apply_conditional_progress_coloring(
            ws = ws,
            base_formula_plan = formula_plan,
            formula_range = formula_range,
            color = color,
            reference_column = ref_column,
            upper_threshold = threshold,
        )

    apply_conditional_current_week_coloring(
        log = log,
        ws = ws,
        color = BG_COLOR_WEEK_HIGHLIGHT,
        formula_range = formula_range_for_week_highlight,
        reference_column = ref_column_for_week_highlight,
        reference_row = ref_row_for_week_highlight,
    )

    for col_letter in col_letters_for_horizon_cols:
        ws.column_dimensions[col_letter].width = 3

    task_name_width = math.ceil(max([
        len(task.name().strip())
        for task in tasks_dict.values()
    ]) * 0.8)
    log.debug("Setting task name column '{}'to width: {}".format(col_letter_task_name, task_name_width))
    ws.column_dimensions[col_letter_task_name].width = task_name_width

    date_row_height = 50
    log.debug("Setting date row height to {}.".format(date_row_height))
    ws.row_dimensions[date_row_num].height = date_row_height

    for (col_start, col_end) in groups_to_hide:
        ws.column_dimensions.group(col_start, col_end, hidden = True)

    for (row_start, row_end) in rows_to_hide:
        log.debug("Hiding rows {} to {} ...".format(row_start, row_end))
        ws.row_dimensions.group(row_start, row_end, hidden = True)

def __tasks_to_excel_meta(log, wb, sheet_meta, version, date_offset, horizon_granularity):
    ws = wb.create_sheet(title = sheet_meta)
    ws.append([
        excelcolumns.COL_META_DATE_OFFSET,
        excelcolumns.COL_META_VERSION_MAJOR,
        excelcolumns.COL_META_VERSION_MINOR,
        excelcolumns.COL_META_VERSION_PATCH,
        excelcolumns.COL_META_HORIZON_GRANULARITY,
    ])
    version_values = version.split(".")
    if not len(version_values) == 3:
        log.error("Cannot parse version information, I skip meta data export and Excel might not work: {}".format(version))
        return

    meta_row = [
        date_offset,
    ]
    meta_row.extend(version_values)
    meta_row.append(str(horizon_granularity).lower())
    ws.append(meta_row)


def tasks_to_excel(
    log, 
    tasks_dict, 
    path = "tmp/export.xlsx", 
    sheet_tasks = "tasks", 
    sheet_meta = "data", 
    version = version, 
    date_offset = datetime.date.today(),
    horizon_granularity = common.GRANULARITY_WEEKS,
    colors = False,
):
    log.debug("Creating Excel export for {} tasks ...".format(len(tasks_dict.values())))
    wb = xl.Workbook()

    __tasks_to_excel_tasks(log, wb, tasks_dict, sheet_tasks, date_offset, horizon_granularity, colors)
    __tasks_to_excel_meta(log, wb, sheet_meta, version, date_offset, horizon_granularity)

    log.debug("Saving Excel workbook to {} ...".format(path))
    try:
        wb.save(path)
    except:
        log.exception("Saving Excel workbook failed. See exception for details.")
        return False

    return True

def __return_with_error(log, col_id):
    log.error("Cannot find column with header '{}' in Excel. Make sure column names are correct.".format(EXCEL_COL_NAMES[col_id]))
    return False, {}    

def tasks_from_excel(log, path = "tmp/input.xlsx"):
    log.debug("Trying to load existing tasks from '{}' ...".format(path))
    success = False
    tasks_dict = {}
    if not os.path.exists(path):
        log.error("Cannot find Excel file '{}' to read from. I skip loading from Excel.".format(path))
        return success, tasks_dict

    try:
        wb = xl.load_workbook(path)
    except:
        log.exception("Loading Excel from {} failed. I skip import.".format(path))
        return success, tasks_dict

    if not excelcolumns.DEFAULT_SHEET_TASKS in wb.sheetnames:
        log.error("There is not worksheet called '{}' in file '{}'. I stop importing meta information.".format(excelcolumns.DEFAULT_SHEET_TASKS, path))
        return success, tasks_dict

    ws = wb[excelcolumns.DEFAULT_SHEET_TASKS]

    # fetch interesting column letters for reading
    col_letter_task_id = common.find_column_for_cell(log, ws, row = 1, cell_value = EXCEL_COL_NAMES[COL_TASK_ID])
    if not col_letter_task_id:
        return __return_with_error(log, col_id = COL_TASK_ID)

    col_letter_task_name = common.find_column_for_cell(log, ws, row = HEADER_ROW, cell_value=EXCEL_COL_NAMES[COL_TASK_NAME])
    col_letter_task_start = common.find_column_for_cell(log, ws, row = HEADER_ROW, cell_value=EXCEL_COL_NAMES[COL_TASK_START])
    col_letter_task_effort = common.find_column_for_cell(log, ws, row = HEADER_ROW, cell_value=EXCEL_COL_NAMES[COL_TASK_EFFORT])
    col_letter_task_progress = common.find_column_for_cell(log, ws, row = HEADER_ROW, cell_value=EXCEL_COL_NAMES[COL_TASK_PROGRESS])

    col_letter_task_costs = common.find_column_for_cell(log, ws, row = HEADER_ROW, cell_value=EXCEL_COL_NAMES[COL_TASK_COSTS])

    stop = False

    # First action in the loop is to increment (there is no do-while loop)
    row_counter = DATA_ROW_START - 1

    while not stop:
        row_counter += 1
        task_id = ws["{}{}".format(col_letter_task_id, row_counter)].value
        stop = task_id == None
        if stop:
            log.debug("Stop reading {} at row {} because no task ID found.".format(path, row_counter))
            break

        log.debug("Reading task info from row {} in {} ...".format(row_counter, path))
        task_name = ws["{}{}".format(col_letter_task_name, row_counter)].value
        task_start = ws["{}{}".format(col_letter_task_start, row_counter)].value
        task_effort = ws["{}{}".format(col_letter_task_effort, row_counter)].value
        task_progress = ws["{}{}".format(col_letter_task_progress, row_counter)].value

        task_costs = 0 if not col_letter_task_costs else ws["{}{}".format(col_letter_task_costs, row_counter)].value

        tasks_dict[task_id] = task.Task(
            id = task_id,
            name = task_name,
            effort = task_effort,
            progress = task_progress,
            start = task_start,
            costs = task_costs,
        )

    success = True
    return success, tasks_dict
