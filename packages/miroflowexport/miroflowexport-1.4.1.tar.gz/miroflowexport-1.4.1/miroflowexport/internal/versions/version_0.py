import os

import openpyxl as xl

from miroflowexport.internal.versions import common
from miroflowexport.internal import task

from miroflowexport.internal.excelcolumns import *

def __return_with_error(log, col_id):
    log.error("Cannot find column with header '{}' in Excel. Make sure column names are correct.".format(EXCEL_COL_NAMES_V0[col_id]))
    return False, {}    

def tasks_from_excel(log, path = "tmp/input.xlsx"):
    log.debug("Trying to load existing tasks from '{}' ...".format(path))
    success = False
    tasks_dict = {}
    if not os.path.exists(path):
        log.error("Cannot find Excel file '{}' to read from. I skip loading from Excel.".format(path))
        return success, tasks_dict

    try:
        wb = xl.load_workbook(path)
    except:
        log.exception("Loading Excel from {} failed. I skip import.".format(path))
        return success, tasks_dict

    ws = wb.active

    # fetch interesting column letters for reading
    col_letter_task_id = common.find_column_for_cell(log, ws, row = 1, cell_value = EXCEL_COL_NAMES_V0[COL_TASK_ID])
    if not col_letter_task_id:
        return __return_with_error(log, col_id = COL_TASK_ID)

    col_letter_task_name = common.find_column_for_cell(log, ws, row = 1, cell_value=EXCEL_COL_NAMES_V0[COL_TASK_NAME])
    col_letter_task_start = common.find_column_for_cell(log, ws, row = 1, cell_value=EXCEL_COL_NAMES_V0[COL_TASK_START])
    col_letter_task_effort = common.find_column_for_cell(log, ws, row = 1, cell_value = EXCEL_COL_NAMES_V0[COL_TASK_EFFORT])
    col_letter_task_progress = common.find_column_for_cell(log, ws, row = 1, cell_value=EXCEL_COL_NAMES_V0[COL_TASK_PROGRESS])

    stop = False
    row_counter = 1

    while not stop:
        row_counter += 1
        task_id = ws["{}{}".format(col_letter_task_id, row_counter)].value
        stop = task_id == None
        if stop:
            log.debug("Stop reading {} at row {} because no task ID found.".format(path, row_counter))
            break

        log.debug("Reading task info from row {} in {} ...".format(row_counter, path))
        task_name = ws["{}{}".format(col_letter_task_name, row_counter)].value
        task_start = ws["{}{}".format(col_letter_task_start, row_counter)].value
        task_effort = ws["{}{}".format(col_letter_task_effort, row_counter)].value
        task_progress = ws["{}{}".format(col_letter_task_progress, row_counter)].value

        tasks_dict[task_id] = task.Task(
            id = task_id,
            name = task_name,
            effort = task_effort,
            progress = task_progress,
            start = task_start,
        )

    success = True
    return success, tasks_dict
