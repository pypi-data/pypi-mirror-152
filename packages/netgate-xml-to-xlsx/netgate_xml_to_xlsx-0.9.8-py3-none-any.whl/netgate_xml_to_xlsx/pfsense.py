"""PfSense class."""
# Copyright © 2022 Appropriate Solutions, Inc. All rights reserved.

import argparse
import os
from pathlib import Path
from typing import cast

from lxml import etree  # nosec
from openpyxl import Workbook
from openpyxl.styles import Border, Font, NamedStyle, PatternFill, Side
from openpyxl.styles.alignment import Alignment

from netgate_xml_to_xlsx.mytypes import Node

from .plugin_tools import discover_plugins
from .plugins.support.elements import sanitize_xml
from .spreadsheet import sheet_footer, sheet_header, write_ss_row


class PfSense:
    """Handle all pfSense parsing and conversion."""

    def __init__(self, args: argparse.Namespace, in_filename: str) -> None:
        """
        Initialize and load XML.

        Technically a bit too much work to do in an init (since it can fail).
        """
        self.args = args
        self.in_file = (in_path := Path(in_filename))
        self.raw_xml: str = ""
        self.parsed_xml: Node = None
        self.workbook: Workbook = Workbook()

        # ss_filename is expected to be overwritten by
        self.ss_output_path = self._get_output_path(in_path)
        self._init_styles()
        self.default_alignment = Alignment(wrap_text=True, vertical="top")

        self.plugins = discover_plugins()

        self._load()

    def _get_output_path(self, in_path: Path) -> Path:
        """Generate output path based on args and in_filename."""
        out_path = cast(Path, self.args.output_dir) / Path(f"{in_path.name}.xlsx")
        return out_path

    def _init_styles(self) -> None:
        """Iniitalized worksheet styles."""
        xlsx_header_font = Font(name="Calibri", size=16, italic=True, bold=True)
        xlsx_body_font = Font(name="Calibri", size=16)
        xlsx_footer_font = Font(name="Calibri", size=12, italic=True)

        body_border = Border(
            bottom=Side(border_style="dotted", color="00000000"),
            top=Side(border_style="dotted", color="00000000"),
            left=Side(border_style="dotted", color="00000000"),
            right=Side(border_style="dotted", color="00000000"),
        )

        alignment = Alignment(wrap_text=True, vertical="top")

        header = NamedStyle(name="header")
        header.alignment = alignment
        header.fill = PatternFill(
            "lightTrellis", fgColor="00339966"
        )  # fgColor="000000FF")  #fgColor="0000FF00")
        header.font = xlsx_header_font
        header.border = Border(
            bottom=Side(border_style="thin", color="00000000"),
            top=Side(border_style="thin", color="00000000"),
            left=Side(border_style="dotted", color="00000000"),
            right=Side(border_style="dotted", color="00000000"),
        )

        normal = NamedStyle(name="normal")
        normal.alignment = alignment
        normal.border = body_border
        normal.fill = PatternFill("solid", fgColor="FFFFFFFF")
        normal.font = xlsx_body_font

        footer = NamedStyle("footer")
        footer.alignment = Alignment(wrap_text=False, vertical="top")
        footer.border = body_border
        normal.fill = PatternFill("solid", fgColor="FFFFFFFF")
        footer.font = xlsx_footer_font

        self.workbook.add_named_style(header)
        self.workbook.add_named_style(normal)
        self.workbook.add_named_style(footer)

    def _sanity_check_root_node(self) -> None:
        """Check for unknown root elements."""
        expected_nodes = set(
            (
                "version,lastchange,revision,aliases,ca,"
                "cert,cron,dhcpd,dhcpdv6,diag,"
                "dnshaper,filter,gateways,hasync,ifgroups,"
                "installedpackages,interfaces,ipsec,nat,notifications,"
                "ntpd,openvpn,ovpnserver,ppps,proxyarp,"
                "rrd,shaper,snmpd,sshdata,staticroutes,"
                "switches,sysctl,syslog,system_groups,system_users,"
                "system,unbound,virtualip,vlans,widgets,"
                "wizardtemp,wol"
            ).split(",")
        )
        unknown = []
        for child in self.parsed_xml:
            if child.tag not in expected_nodes:
                unknown.append(child.tag)
        if len(unknown):
            print(f"""Unknown root node(s): {",".join(unknown)}""")

    def _load(self) -> None:
        """Load and parse Netgate xml firewall configuration.

        Return pfsense keys.
        """
        self.raw_xml = self.in_file.read_text(encoding="utf-8")
        self.parsed_xml = etree.XML(self.raw_xml)
        self._sanity_check_root_node()

    def _write_sheet(
        self,
        *,
        sheet_name: str,
        header_row: list[str],
        column_widths: list[int],
        rows: list[list],
    ) -> None:
        sheet = self.workbook.create_sheet(sheet_name)
        sheet_header(sheet, header_row, column_widths)

        # Define starting row num in case there are no rows to display.
        row_num = 2
        for row_num, row in enumerate(rows, start=row_num):
            write_ss_row(sheet, row, row_num)
        sheet_footer(sheet, row_num)

    def sanitize(self, plugins_to_run) -> None:
        """
        Sanitize the raw XML and save as original filename + '-sanitized'.

        The Netgate configuration file XML is well ordered and thus searchable via regex.

        Parse the XML into a tree and call the individual plugin sanitizers.
        Call the individual plugin sanitizers.

        Args:
            plugins_to_run: List of active plugin names.

        """
        # Run generic sanitize.
        self.raw_xml = sanitize_xml(self.raw_xml)

        # Parse xml for plugin sanitizing.
        self.parsed_xml = etree.XML(self.raw_xml)

        for plugin_name in plugins_to_run:
            plugin = self.plugins[plugin_name]
            plugin.sanitize(self.parsed_xml)

        # Pretty format XML.
        self.raw_xml = etree.tostring(self.parsed_xml, pretty_print=True).decode("utf8")

        # Save sanitized XML
        parts = os.path.splitext(self.in_file)
        if len(parts) == 1:
            out_path = Path(f"{parts[0]}-sanitized")
        else:
            out_path = Path(f"{parts[0]}-sanitized{parts[1]}")
        out_path.write_text(f"{self.raw_xml}", encoding="utf-8")
        print(f"Sanitized file written: {out_path}.")

        # Delete the unsanitized file.
        self.in_file.unlink()
        print(f"Deleted original file: {self.in_file}.")

    def run(self, plugin_name: str) -> None:
        """
        Run specific plugin and write sheet(s).

        Plugins yield a SheetData object.
        Continue iterating if it is None.
        """
        plugin = self.plugins[plugin_name]
        for sheet_data in plugin.run(self.parsed_xml):
            if sheet_data is None or not sheet_data.data_rows:
                continue

            self._write_sheet(
                sheet_name=sheet_data.sheet_name,
                header_row=sheet_data.header_row,
                column_widths=sheet_data.column_widths,
                rows=sheet_data.data_rows,
            )

    def save(self) -> None:
        """Delete empty first sheet and then save Workbook."""
        sheets = self.workbook.sheetnames
        del self.workbook[sheets[0]]
        out_path = self.ss_output_path
        self.workbook.save(out_path)
