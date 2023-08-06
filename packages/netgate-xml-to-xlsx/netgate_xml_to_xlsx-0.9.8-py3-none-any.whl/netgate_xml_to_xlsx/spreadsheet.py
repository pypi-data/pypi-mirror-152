"""Stand-alone spreadsheet functions."""
# Copyright © 2022 Appropriate Solutions, Inc. All rights reserved.

import datetime

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def sheet_header(sheet: Worksheet, columns: list, column_widths: list[int]) -> None:
    """Write header row then set the column widths."""
    write_ss_row(sheet, columns, 1, "header")

    for column_number, width in enumerate(column_widths, start=1):
        column_letter = get_column_letter(column_number)
        sheet.column_dimensions[column_letter].width = width


def write_ss_row(
    sheet: Worksheet, row: list, row_num: int, style_name: str = "normal"
) -> None:
    """
    Write a row into the spreadsheet.

    Args:
        row: A list of values to write into the row.

        row_increment: Number of rows to increment in spreadsheet before writing.

        style_name: Named XLSX style.

    Always increment the row before writing.

    """
    for column_number, value in enumerate(row, start=1):
        column_letter = get_column_letter(column_number)
        coordinate = f"{column_letter}{row_num}"
        sheet[coordinate] = value
        sheet[coordinate].style = style_name


def sheet_footer(sheet: Worksheet, row_number: int) -> None:
    """Write footer information on each sheet."""
    now = datetime.datetime.now().strftime("%Y-%m-%dT%H:%M")
    run_date = f"Run date: {now}"

    write_ss_row(sheet, [run_date], row_number + 1, style_name="footer")
