import pathlib

from openpyxl import load_workbook


def load_exce_all_sheet(path: pathlib.Path):
    suite_list = []
    wb = load_workbook(path, read_only=True)
    for ws in wb.worksheets:
        _suite = {"info": {}, "case_list": []}
        _case = _suite["case_list"]
        new_case = {}
        top_keys = tuple(
            [
                i
                for i in list(
                    ws.iter_rows(
                        max_row=1,
                        values_only=True,
                    )
                )[0]
                if i is not None
            ]
        )
        # print(top_keys)
        for (
            case_id,
            step_id,
            step_name,
            keyword,
            *args,
            result,
            summary,
        ) in ws.iter_rows(min_row=2, values_only=True):
            if case_id == 0:
                _suite["info"][keyword] = args[0]
                continue
            elif step_id == 0:
                _case.append(new_case)
                new_case = {"info": {}, "steps": []}
                new_case["info"][keyword] = args[0]
            else:

                new_case["steps"].append(
                    dict(
                        zip(
                            top_keys,
                            (
                                case_id,
                                step_id,
                                step_name,
                                keyword,
                                [_arg for _arg in args if _arg is not None],
                                result,
                                summary,
                            ),
                        )
                    )
                )
        _case.append(new_case)
        _case.remove({})
        yield _suite

    wb.close()
