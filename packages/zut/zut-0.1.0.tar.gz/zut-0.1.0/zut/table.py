import csv, re, os, sys
from datetime import datetime, timezone
from tabulate import tabulate
from openpyxl import Workbook, DEFUSEDXML
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import get_column_letter

if not DEFUSEDXML:
    raise ValueError("By default openpyxl does not guard against quadratic blowup or billion laughs xml attacks. To guard against these attacks install defusedxml.")

# Register CSV dialect for French version of Excel
class ExcelFr(csv.excel):
    delimiter = ";"

csv.register_dialect('excel-fr', ExcelFr())

class OpenTable:
    def __init__(self, target, header=None, format=None, **options):
        if format and (not target or isinstance(target, str)):
            raise ValueError("format can only be given when target is not a string")

        if not target:
            self.writer = NoneTableWriter()
            return

        if not isinstance(target, str):
            if format == "xlsx":
                self.writer = XlsxTableWriter(header, target=target, **options)
            elif format == "csv":
                self.writer = CsvTableWriter(header, target=target, **options)
            else:
                self.writer = TabulateTableWriter(header, target=target, **options)
            return

        if target in TabulateTableWriter.targetstrings:
            self.writer = TabulateTableWriter(header, target=target, **options)
            return
        
        m = re.match(r'^(.+)(\.[^\.#]+)(#[^\.#]+)?$', target)
        if not m:
            raise ValueError("Invalid target: %s" % target)

        main = m.group(1)
        ext = m.group(2).lower()
        tablename = m.group(3)

        if ext == '.xlsx':
            self.writer = XlsxTableWriter(header, target=main + '.xlsx', tablename=tablename, **options)
            return

        if tablename:
            raise ValueError("Invalid target: %s (table name can only be specified for Excel files)" % target)

        if ext == '.csv':
            self.writer = CsvTableWriter(header, target=main + '.csv', **options)
            return

        raise ValueError("Invalid target: %s (invalid extension)" % target)
    
    def __enter__(self):
        self.writer.__enter__()
        return self.writer

    def __exit__(self, *args):
        self.writer.__exit__(*args)


class AbstractTableWriter:
    def __init__(self, header, target, options={}):
        self.header = header
        self.row_count = 0
        self.target = target
        self.options = options

    def append(self, row, is_header=False):
        if is_header:
            if self.row_count > 0:
                raise ValueError("Cannot append header: rows already appended")
        else:
            self.row_count += 1

    def __enter__(self):
        # Create parent directories if necessary
        if self.target and isinstance(self.target, str):
            dirpath = os.path.dirname(self.target)
            if dirpath:
                os.makedirs(dirpath, exist_ok=True)
    
    def __exit__(self, *args):
        pass


class NoneTableWriter(AbstractTableWriter):
    def __init__(self):
        super().__init__(header=None, target=None)


class TabulateTableWriter(AbstractTableWriter):
    targetstrings = ["stdout", "stderr", "tabulate"]

    def __init__(self, header, target, label=None, ljust=None, border_top=False, border_bottom=False):
        if isinstance(target, str):
            if target == "stderr":
                target = sys.stderr
            else:
                target = sys.stdout
        super().__init__(header, target, options={"label": label, "ljust": ljust, "border_top": border_top, "border_bottom": border_bottom})

    def __enter__(self):
        super().__enter__()
        self.tabulate_rows = []

    def __exit__(self, *args):
        if self.options["label"]:
            label = self.options["label"]
            if self.options["ljust"]:
                label = label.ljust(self.options["ljust"])
        else:
            label = ""

        if not self.tabulate_rows:
            if label:
                print(label, file=self.target)
            return

        text = tabulate(self.tabulate_rows, headers=self.header)
        lines = text.splitlines()

        if self.options["ljust"]:
            ljust = " " * self.options["ljust"]
        elif label:
            ljust = " " * len(label)
        else:
            ljust = ""

        if self.options["border_top"]:
            print(ljust + lines[1], file=self.target)

        print((label if label else ljust) + lines[0], file=self.target)
        for line in lines[1:]:
            print(ljust + line, file=self.target)

        if self.options["border_bottom"]:
            print(ljust + lines[1], file=self.target)

    def append(self, row, is_header=False):
        super().append(row, is_header=is_header)
        if not is_header:
            self.tabulate_rows.append(row)


class CsvTableWriter(AbstractTableWriter):
    def __init__(self, header, target, dialect=None):
        if not dialect:
            dialect = os.environ.get('CSV_DIALECT', 'excel-fr' if 'LANG' in os.environ and os.environ['LANG'].startswith('fr') else 'excel')
        super().__init__(header, target, options={"dialect": dialect})
    
    def __enter__(self):
        super().__enter__()

        if isinstance(self.target, str):
            self.csv_file = open(self.target, 'w', newline='')
        else:
            self.csv_file = self.target
        self.csv_writer = csv.writer(self.csv_file, dialect=self.options["dialect"])
        
        if self.header:
            self.csv_writer.writerow(self.header)
    
    def __exit__(self, *args):
        if isinstance(self.target, str):
            self.csv_file.close()

    def append(self, row, is_header=False):
        super().append(row, is_header=is_header)
        self.csv_writer.writerow(row)


class XlsxTableWriter(AbstractTableWriter):
    def __init__(self, header, target, tablename=None, timezone=None):
        super().__init__(header, target, options={"timezone": timezone})
        if tablename:
            if not header:
                raise ValueError("Cannot create tab without headers")
            if tablename.startswith('#'):
                tablename = tablename[1:]
            self.xlsx_tablename = tablename
        elif self.header:
            self.xlsx_tablename = "Data"
        else:
            self.xlsx_tablename = None
        
    def __enter__(self):
        super().__enter__()

        self.xlsx_workbook = Workbook(write_only=True)
        self.xlsx_worksheet = self.xlsx_workbook.create_sheet(self.xlsx_tablename if self.xlsx_tablename else "Data")

        if self.header:
            self.xlsx_worksheet.append(self.header)

    def __exit__(self, *args):
        if self.xlsx_tablename:
            max_column_letter = get_column_letter(len(self.header))
            max_row_number = 1 + self.row_count

            # In write-only mode, table columns must be added manually
            table = Table(displayName=self.xlsx_tablename, ref=f"A1:{max_column_letter}{max_row_number}")
            table._initialise_columns()
            for column, value in zip(table.tableColumns, self.header):
                column.name = value

            # Add table styling
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style

            setattr(self.xlsx_worksheet, "_get_cell", None) # remove write-only warning
            self.xlsx_worksheet.add_table(table)
            delattr(self.xlsx_worksheet, "_get_cell")

        self.xlsx_workbook.save(self.target)
        self.xlsx_workbook.close()

    def append(self, row, is_header=False):
        super().append(row, is_header=is_header)

        for i, value in enumerate(row):
            if isinstance(value, datetime) and value.tzinfo:
                if value.tzinfo:
                    # Excel does not support timezones in datetimes
                    if self.options["timezone"]:
                        value = value.astimezone(self.options["timezone"])
                    row[i] = value.replace(tzinfo=None)

        self.xlsx_worksheet.append(row)
